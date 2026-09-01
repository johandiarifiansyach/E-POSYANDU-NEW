"""Train an offline, anthropometry-first model from a cohort workbook.

The Sigizi cohort export stores one child in a seven-row block: the first row
contains monthly weights and the next row contains monthly lengths/heights.
This script converts those blocks into chronological observations and trains
screening models for the status at the next available measurement.  It is an
offline candidate only; the deterministic WHO calculation remains the source
of truth and the resulting artifacts are never loaded by production
automatically.

Exclusive breastfeeding is deliberately excluded from this baseline.  The
training target is a current WHO classification at the next observation, not
a diagnosis or a validated clinical outcome.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import joblib
import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import HistGradientBoostingClassifier
from sklearn.impute import SimpleImputer
from sklearn.metrics import (
    average_precision_score,
    balanced_accuracy_score,
    classification_report,
    roc_auc_score,
)
from sklearn.model_selection import GroupShuffleSplit
from sklearn.pipeline import Pipeline


TRAINING_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(TRAINING_DIR.parent))
from analysis_service.who import assess_item  # noqa: E402


RANDOM_STATE = 20260831
MODEL_VERSION = "growth-trend-hgb-v1-candidate"
MONTHS = ("Jan", "Feb", "Mrt", "Apr", "Mei", "Jun", "Jul", "Ags")

FEATURE_COLUMNS = [
    "age_months",
    "sex_l",
    "weight_kg",
    "height_cm",
    "weight_delta_1",
    "height_delta_1",
    "weight_velocity_per_month",
    "height_velocity_per_month",
    "weight_delta_window",
    "height_delta_window",
    "months_since_previous",
    "trend_points",
    "weight_not_rising",
    "height_not_rising",
]

TARGET_DEFINITIONS = {
    "next_underweight_problem": {
        "statusField": "bbu_status",
        "positive": ["Berat Kurang", "Berat Sangat Kurang"],
    },
    "next_stunting_problem": {
        "statusField": "tbu_status",
        "positive": ["Pendek", "Sangat Pendek"],
    },
    "next_wasting_problem": {
        "statusField": "bbtb_status",
        "positive": ["Gizi Kurang", "Gizi Buruk"],
    },
    "next_overweight_problem": {
        "statusField": "bbtb_status",
        "positive": ["Gizi Lebih", "Obesitas"],
    },
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _number(value: Any) -> bool:
    return isinstance(value, (int, float, np.number)) and not isinstance(value, bool) and math.isfinite(float(value))


def normalize_weight(value: Any) -> tuple[float | None, str]:
    """Normalize common cohort entry mistakes while retaining a quality flag."""

    if not _number(value) or float(value) <= 0:
        return None, "missing"
    number = float(value)
    if number > 200:
        return number / 1000.0, "grams_to_kg"
    if number > 60:
        return number / 10.0, "decimal_shift"
    if 0.1 <= number <= 60:
        return number, "ok"
    return None, "invalid"


def normalize_height(value: Any) -> tuple[float | None, str]:
    """Normalize decimal-shifted height entries from the cohort export."""

    if not _number(value) or float(value) <= 0:
        return None, "missing"
    number = float(value)
    if number >= 1000:
        return number / 100.0, "decimal_shift_two"
    if number > 220:
        return number / 10.0, "decimal_shift_one"
    if number < 20:
        return number * 10.0, "decimal_shift_low"
    if 10 <= number <= 220:
        return number, "ok"
    return None, "invalid"


def _status_for_measurement(weight: float, height: float, age_months: int, sex: str) -> dict[str, Any]:
    # The cohort does not carry measurement method.  Use the usual method by
    # age so the WHO length/height correction remains deterministic.
    method = "Terlentang" if age_months <= 24 else "Berdiri"
    return assess_item(
        {
            "weight_kg": weight,
            "height_cm": height,
            "age_months": age_months,
            "sex": sex,
            "measurement_method": method,
        }
    )


def parse_cohort(source: Path, year: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    # ``ReadOnlyWorksheet.cell`` reparses the XML stream for every random
    # access, which makes a large cohort appear to hang.  Materialize the
    # small 16-column sheet once and then walk the seven-row blocks.
    rows = list(sheet.iter_rows(values_only=True))
    quality = Counter()
    observations: list[dict[str, Any]] = []
    children = 0

    for base_index in range(5, len(rows), 7):
        child_row = rows[base_index]
        height_row = rows[base_index + 1] if base_index + 1 < len(rows) else ()
        child_id = child_row[0] if len(child_row) > 0 else None
        if child_id is None:
            continue
        children += 1
        quality["children"] = children
        birth_value = child_row[1] if len(child_row) > 1 else None
        birth_date = pd.to_datetime(birth_value, errors="coerce")
        sex = str(child_row[2] if len(child_row) > 2 else "").strip().upper()
        if pd.isna(birth_date) or sex not in {"L", "P"}:
            quality["children_invalid_identity"] += 1
            continue

        for month_number, month_name in enumerate(MONTHS, 1):
            # Excel column E (zero-based tuple index 4) is January.
            column_index = month_number + 3
            weight_raw = child_row[column_index] if len(child_row) > column_index else None
            height_raw = height_row[column_index] if len(height_row) > column_index else None
            weight, weight_flag = normalize_weight(weight_raw)
            height, height_flag = normalize_height(height_raw)
            quality[f"weight_{weight_flag}"] += 1
            quality[f"height_{height_flag}"] += 1
            if weight is None or height is None:
                if weight is not None or height is not None:
                    quality["partial_pairs"] += 1
                continue

            quality["valid_pairs"] += 1
            measured = pd.Timestamp(year=year, month=month_number, day=15)
            age_float = (measured - birth_date).days / 30.4375
            age_months = math.floor(age_float)
            if not 0 <= age_months <= 60:
                quality["pairs_age_outside_0_60"] += 1
                continue
            status = _status_for_measurement(weight, height, age_months, sex)
            observations.append(
                {
                    "child_id": str(child_id),
                    "month_number": month_number,
                    "month": month_name,
                    "measurement_date": measured.strftime("%Y-%m-%d"),
                    "age_months": age_months,
                    "sex_l": 1.0 if sex == "L" else 0.0,
                    "weight_kg": weight,
                    "height_cm": height,
                    "statuses": {
                        key: status[key]
                        for key in ("bbu_status", "tbu_status", "bbtb_status")
                    },
                }
            )

    quality["observations"] = len(observations)
    quality["children_with_observations"] = len({item["child_id"] for item in observations})
    quality["jan_measurements"] = sum(item["month_number"] == 1 for item in observations)
    quality["feb_aug_measurements"] = sum(2 <= item["month_number"] <= 8 for item in observations)
    return observations, dict(quality)


def build_trend_examples(observations: list[dict[str, Any]]) -> pd.DataFrame:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for observation in observations:
        grouped.setdefault(observation["child_id"], []).append(observation)

    examples: list[dict[str, Any]] = []
    for child_id, records in grouped.items():
        records.sort(key=lambda item: item["month_number"])
        # A trend example needs a previous point for its slope and a next
        # point for the target.  The child id is used only for grouping.
        for index in range(1, len(records) - 1):
            previous = records[index - 1]
            current = records[index]
            following = records[index + 1]
            elapsed = max(1, current["month_number"] - previous["month_number"])
            window = records[max(0, index - 2) : index + 1]
            window_elapsed = max(1, current["month_number"] - window[0]["month_number"])
            weight_delta = current["weight_kg"] - previous["weight_kg"]
            height_delta = current["height_cm"] - previous["height_cm"]
            examples.append(
                {
                    "child_id": child_id,
                    "age_months": current["age_months"],
                    "sex_l": current["sex_l"],
                    "weight_kg": current["weight_kg"],
                    "height_cm": current["height_cm"],
                    "weight_delta_1": weight_delta,
                    "height_delta_1": height_delta,
                    "weight_velocity_per_month": weight_delta / elapsed,
                    "height_velocity_per_month": height_delta / elapsed,
                    "weight_delta_window": current["weight_kg"] - window[0]["weight_kg"],
                    "height_delta_window": current["height_cm"] - window[0]["height_cm"],
                    "months_since_previous": elapsed,
                    "trend_points": len(window),
                    "weight_not_rising": 1.0 if weight_delta <= 0 else 0.0,
                    "height_not_rising": 1.0 if height_delta <= 0 else 0.0,
                    "next_underweight_problem": following["statuses"]["bbu_status"]
                    in {"Berat Kurang", "Berat Sangat Kurang"},
                    "next_stunting_problem": following["statuses"]["tbu_status"]
                    in {"Pendek", "Sangat Pendek"},
                    "next_wasting_problem": following["statuses"]["bbtb_status"]
                    in {"Gizi Kurang", "Gizi Buruk"},
                    "next_overweight_problem": following["statuses"]["bbtb_status"]
                    in {"Gizi Lebih", "Obesitas"},
                }
            )
    return pd.DataFrame(examples)


def _split_examples(examples: pd.DataFrame, test_size: float) -> tuple[np.ndarray, np.ndarray, int]:
    for offset in range(20):
        random_state = RANDOM_STATE + offset
        splitter = GroupShuffleSplit(n_splits=1, test_size=test_size, random_state=random_state)
        train_idx, test_idx = next(splitter.split(examples, groups=examples["child_id"]))
        valid = True
        for target in TARGET_DEFINITIONS:
            train_values = examples.iloc[train_idx][target].astype(int)
            test_values = examples.iloc[test_idx][target].astype(int)
            if train_values.nunique() < 2 or test_values.nunique() < 2:
                valid = False
                break
        if valid:
            return train_idx, test_idx, random_state
    raise ValueError("Tidak menemukan split per anak dengan dua kelas untuk semua target.")


def _metrics(y_true: np.ndarray, probabilities: np.ndarray) -> dict[str, Any]:
    predictions = (probabilities >= 0.5).astype(int)
    result: dict[str, Any] = {
        "support": int(len(y_true)),
        "positive": int(y_true.sum()),
        "negative": int((1 - y_true).sum()),
        "balancedAccuracy": round(float(balanced_accuracy_score(y_true, predictions)), 4),
        "classificationReport": classification_report(
            y_true,
            predictions,
            labels=[0, 1],
            target_names=["normal_or_other", "problem"],
            output_dict=True,
            zero_division=0,
        ),
    }
    result["rocAuc"] = round(float(roc_auc_score(y_true, probabilities)), 4)
    result["averagePrecision"] = round(float(average_precision_score(y_true, probabilities)), 4)
    return result


def train(args: argparse.Namespace) -> dict[str, Any]:
    source = Path(args.xlsx).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    observations, quality = parse_cohort(source, args.year)
    examples = build_trend_examples(observations)
    if len(examples) < 100:
        raise ValueError("Contoh tren terlalu sedikit setelah pembersihan.")
    train_idx, test_idx, split_seed = _split_examples(examples, args.test_size)

    models: dict[str, Any] = {}
    metrics: dict[str, Any] = {}
    for name, definition in TARGET_DEFINITIONS.items():
        x_train = examples.iloc[train_idx][FEATURE_COLUMNS]
        y_train = examples.iloc[train_idx][name].astype(int)
        x_test = examples.iloc[test_idx][FEATURE_COLUMNS]
        y_test = examples.iloc[test_idx][name].astype(int)
        model = Pipeline(
            [
                ("imputer", SimpleImputer(strategy="median", add_indicator=True)),
                (
                    "classifier",
                    HistGradientBoostingClassifier(
                        max_iter=180,
                        learning_rate=0.06,
                        max_leaf_nodes=15,
                        min_samples_leaf=24,
                        l2_regularization=1.0,
                        class_weight="balanced",
                        random_state=RANDOM_STATE,
                    ),
                ),
            ]
        )
        model.fit(x_train, y_train)
        probabilities = model.predict_proba(x_test)[:, 1]
        models[name] = model
        metrics[name] = {
            "target": definition,
            "train": {"support": int(len(y_train)), "positive": int(y_train.sum())},
            "test": _metrics(y_test.to_numpy(), probabilities),
        }

    metadata = {
        "modelVersion": MODEL_VERSION,
        "trainedAt": datetime.now(timezone.utc).isoformat(),
        "featureColumns": FEATURE_COLUMNS,
        "targetDefinitions": TARGET_DEFINITIONS,
        "featurePolicy": {
            "anthropometryFirst": True,
            "excludedFeatures": ["exclusive_breastfeeding"],
            "primarySignals": [
                "weight_kg",
                "height_cm",
                "age_months",
                "weight_delta_1",
                "height_delta_1",
                "weight_velocity_per_month",
                "height_velocity_per_month",
            ],
        },
        "labelPolicy": {
            "meaning": "Status WHO pada pengukuran berikutnya yang tersedia.",
            "currentStatusIncludedAsFeature": False,
            "notForDiagnosis": True,
            "notValidatedAsFutureClinicalRisk": True,
        },
        "dataset": {
            "file": str(source),
            "sha256": sha256_file(source),
            "year": args.year,
            "children": quality.get("children", 0),
            "observations": len(observations),
            "trendExamples": len(examples),
        },
        "dataQuality": quality,
        "cohortInterpretation": {
            "layout": "7-row child block; first row monthly weight, second row monthly height/length",
            "measurementMonths": list(MONTHS),
            "monthDateAssumption": "Tanggal 15 digunakan sebagai titik tengah bulan untuk menghitung usia selesai dalam bulan.",
            "zeroAndTidak": "Nilai 0 atau TIDAK diperlakukan sebagai pengukuran yang tidak tersedia.",
            "methodAssumption": "Terlentang untuk usia <=24 bulan, Berdiri untuk usia >24 bulan karena kolom cara ukur tidak tersedia.",
        },
        "split": {
            "strategy": "GroupShuffleSplit",
            "groupKey": "child_id (No anonim)",
            "testSize": args.test_size,
            "randomState": split_seed,
        },
    }
    artifact = {"models": models, "metadata": metadata}
    joblib.dump(artifact, output_dir / "growth_trend_models.joblib", compress=3)
    (output_dir / "trend_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    (output_dir / "trend_metrics.json").write_text(
        json.dumps(metrics, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    return {"outputDir": str(output_dir), "metadata": metadata, "metrics": metrics}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="Anonymised cohort workbook export")
    parser.add_argument("--output-dir", required=True, help="Directory for local trend model artifacts")
    parser.add_argument("--year", type=int, default=2026, help="Year represented by the monthly columns")
    parser.add_argument("--test-size", type=float, default=0.2)
    return parser.parse_args()


if __name__ == "__main__":
    result = train(parse_args())
    print(json.dumps({"outputDir": result["outputDir"], "metrics": result["metrics"]}, ensure_ascii=False, indent=2, default=str))
